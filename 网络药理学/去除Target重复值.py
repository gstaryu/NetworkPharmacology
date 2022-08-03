"""去除 所有药品3（合并表1表2）.xlsx 中的重复值和空行，并保存在 所有药品4（去除重复值）.xlsx 中"""
import openpyxl

wb3 = openpyxl.load_workbook(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据\所有药品3（合并表1表2）.xlsx")  # 打开表3
ws3 = wb3.worksheets[0]  # 指定工作表

wb4 = openpyxl.Workbook()
ws4 = wb4.active

ws4[f"A{1}"] = 'Mol ID'
ws4[f"B{1}"] = 'Molecule name'
ws4[f"C{1}"] = 'Target name'
ws4[f"D{1}"] = 'Gene Symbol'

target = []
count = 2
for row in range(2, ws3.max_row + 1):
    if ws3[f"D{row}"].value is not None:
        if ws3[f"D{row}"].value not in target:
            ws4[f"A{count}"].value = ws3[f"B{row}"].value
            ws4[f"B{count}"].value = ws3[f"C{row}"].value
            ws4[f"C{count}"].value = ws3[f"D{row}"].value
            ws4[f"D{count}"].value = ws3[f"E{row}"].value
            target.append(str(ws3[f"D{row}"].value).strip())
            count += 1
wb4.save(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据\所有药品4（去除重复值）.xlsx")
