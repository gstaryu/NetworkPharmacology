"""合并 所有药品（蛋白质名称转基因名称）.xlsx 和 所有药品2（蛋白质名称转基因名称）.xlsx 两表数据，
以表1为主，表1的Gene Name为空且表2的Gene Name不为空的，将表2的Gene Name填入表1的Gene Name中，将结果保存在 所有药品3（合并表1表2）.xlsx 中"""
import openpyxl

wb1 = openpyxl.load_workbook(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据\所有药品（蛋白质名称转基因名称）.xlsx")  # 打开表1
ws1 = wb1.worksheets[0]  # 指定工作表
wb2 = openpyxl.load_workbook(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据\所有药品2（蛋白质名称转基因名称）.xlsx")  # 打开表2
ws2 = wb2.worksheets[0]  # 指定工作表

wb3 = openpyxl.Workbook()
ws3 = wb3.active

ws3[f"A{1}"] = 'Medicine'
ws3[f"B{1}"] = 'Mol ID'
ws3[f"C{1}"] = 'Molecule name'
ws3[f"D{1}"] = 'Target name'
ws3[f"E{1}"] = 'Gene Symbol'

count = 0
for row in range(2, ws1.max_row + 1):
    ws3[f"A{row}"].value = ws1[f"A{row}"].value
    ws3[f"B{row}"].value = ws1[f"B{row}"].value
    ws3[f"C{row}"].value = ws1[f"C{row}"].value
    ws3[f"D{row}"].value = ws1[f"D{row}"].value
    if ws1[f"E{row}"].value is not None:
        ws3[f"E{row}"].value = ws1[f"E{row}"].value
        count += 1
    elif ws1[f"E{row}"].value is None and ws2[f"E{row}"].value is not None:
        ws3[f"E{row}"].value = ws2[f"E{row}"].value
        count += 1

print(f"匹配率：{count / (ws1.max_row - 1) * 100}%")
wb3.save(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据\所有药品3（合并表1表2）.xlsx")
