"""将所有药品中的蛋白质名称（Target name）转换成Gene Symbol，并写入 所有药品（蛋白质名称转基因名称）.xlsx 文件中，该条件下有72%的匹配率"""
import openpyxl
import re
import winsound

wbMedicine = openpyxl.load_workbook(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据" + "\\" + "所有药品.xlsx")  # 打开药品数据文件
wsMedicine = wbMedicine.active
wbUniport = openpyxl.load_workbook(r"C:\Users\guo\Desktop\uniprot-Human.xlsx")  # 打开uniprot文件
wsUniport = wbUniport.active

wbTransform = openpyxl.Workbook()  # 新建一个文件
wsTransform = wbTransform.active
wsTransform[f"A{1}"] = 'Medicine'
wsTransform[f"B{1}"] = 'Mol ID'
wsTransform[f"C{1}"] = 'Molecule name'
wsTransform[f"D{1}"] = 'Target name'
wsTransform[f"E{1}"] = 'Gene Symbol'

# medicineProteinList = []  # 药品蛋白质列表
# uniportProteinList = []  # uniport蛋白质列表
# uniportGeneList = []  # uniport基因列表
#
# for row in range(2, wsUniport.max_row + 1):
#     proteinName = str(wsUniport[f"C{row}"].value)  # 取出单元格中的值
#     tempList = []  # 新建一个临时列表，用于存放该单元格中蛋白质的每个名称
#     tempList.extend(re.findall("[\s\S]*\[.*?]\s\w*", proteinName))  # 匹配特殊情况并合并列表，re.findall返回一个列表
#     temp = re.sub('\[(.*?)]', '', proteinName)  # 去除所有方括号及方括号中的内容
#     tempList.extend(re.findall(r"[(](.*?)[)]", temp))  # 匹配所有小括号中的内容（及蛋白质的别名），并合并列表
#     temp2 = re.sub(r"[(](.*?)[)]", '', temp)  # 去除所有小括号及小括号中的内容
#     tempList.append(temp2)  # 将剩下的内容加入列表
#
#     geneName = str(wsUniport[f"E{row}"].value).split(';')
#     uniportGeneList.append(geneName)

for row in range(2, wsMedicine.max_row + 1):
    medicineProtein = str(wsMedicine[f"D{row}"].value)
    wsTransform[f"A{row}"].value = wsMedicine[f"A{row}"].value
    wsTransform[f"B{row}"].value = wsMedicine[f"B{row}"].value
    wsTransform[f"C{row}"].value = wsMedicine[f"C{row}"].value
    wsTransform[f"D{row}"].value = wsMedicine[f"D{row}"].value
    flag = 0  # 是否匹配成功的标记
    for uniport_row in range(2, wsUniport.max_row + 1):
        if flag:
            break
        proteinName = str(wsUniport[f"C{uniport_row}"].value)  # 取出单元格中的值
        tempList = []  # 新建一个临时列表，用于存放该单元格中蛋白质的每个名称
        tempList.extend(re.findall("[\s\S]*\[.*?]\s\w*", proteinName))  # 匹配特殊情况并合并列表，re.findall返回一个列表
        temp = re.sub('\[(.*?)]', '', proteinName)  # 去除所有方括号及方括号中的内容
        tempList.extend(re.findall(r"[(](.*?)[)]", temp))  # 匹配所有小括号中的内容（及蛋白质的别名），并合并列表
        temp2 = re.sub(r"[(](.*?)[)]", '', temp)  # 去除所有小括号及小括号中的内容
        tempList.append(temp2)  # 将剩下的内容加入列表
        for protein in tempList:
            if medicineProtein.strip() == protein.strip():
                wsTransform[f"E{row}"].value = wsUniport[f"E{uniport_row}"].value
                flag = 1
                break
wbTransform.save(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据\所有药品（蛋白质名称转基因名称）.xlsx")
winsound.Beep(1000, 440)  # 代码运行结束后发出警报提示，持续1s
