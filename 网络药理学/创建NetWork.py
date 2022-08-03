import openpyxl

s = "甘草、茯苓、白术、半夏、陈皮、鸡内金、党参、麦芽、砂仁、太子参、枳壳、木香、神曲"
s = s.split("、")
wb = openpyxl.Workbook()
ws = wb.active
ws[f'A{1}'].value = "ID"
ws[f'B{1}'].value = "Medicine"
ws[f'C{1}'].value = "Mol ID"
ws[f'D{1}'].value = "Gene Symbol"

count = 2
for file in s:
    wb1 = openpyxl.load_workbook(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据" + '\\' + file + ".xlsx")  # 打开文件
    sheet = wb1.worksheets[2]  # 指定工作表
    for row in range(2, sheet.max_row + 1):
        ws.cell(row=count, column=2).value = str(file)
        ws.cell(row=count, column=3).value = sheet.cell(row=row, column=1).value
        ws.cell(row=count, column=4).value = sheet.cell(row=row, column=4).value
        count += 1
wb.save(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据" + '\\' + "NetWork.xlsx")
wb = openpyxl.load_workbook(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据" + '\\' + "NetWork.xlsx")
ws = wb.worksheets[0]
common = {'MOL001755': 'A', 'MOL005828': 'B', 'MOL007514': 'C', 'MOL000211': 'D', 'MOL003896': 'E', 'MOL007180': 'F',
          'MOL000006': 'G', 'MOL004328': 'H', 'MOL000449': 'I', 'MOL000359': 'J', 'MOL000358': 'K'}
medicine = {'甘草': 'GC', '茯苓': 'FL', '白术': 'BZ', '半夏': 'BX', '陈皮': 'CP', '鸡内金': 'JNJ', '党参': 'DS', '麦芽': 'MY',
            '砂仁': 'SR', '太子参': 'TZS', '枳壳': 'ZK', '木香': 'MX'}
num = 0
for row in range(2, ws.max_row + 1):
    if ws.cell(row=row, column=2).value != ws.cell(row=row - 1, column=2).value:
        print(num)
        num = 0
    if ws.cell(row=row, column=3).value in common:
        ws.cell(row=row, column=1).value = common[ws.cell(row=row, column=3).value]
    else:
        if ws.cell(row=row, column=2).value in medicine:
            if ws.cell(row=row, column=3).value != ws.cell(row=row - 1, column=3).value:
                num += 1
            ws.cell(row=row, column=1).value = f"{medicine[ws.cell(row=row, column=2).value]}{num}"
wb.save(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据" + '\\' + "NetWork.xlsx")
