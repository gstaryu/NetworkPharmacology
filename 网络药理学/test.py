import openpyxl

wb1 = openpyxl.load_workbook(r"D:\胃癌\创新学习\论文及数据\匹配到的基因.xlsx")
ws1 = wb1.worksheets[0]

print(ws1.max_row)
for row in range(1,ws1.max_row+1):
    print(f'{ws1[f"A{row}"].value}',end='、')