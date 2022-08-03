"""将每味药的蛋白质名称（Target name）转换成Gene Symbol，写入sheet3（全部信息）中"""
import openpyxl

wb4 = openpyxl.load_workbook(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据\所有药品4（去除重复值）.xlsx")
ws4 = wb4.worksheets[0]
targetToGene = {}
for row in range(2, ws4.max_row + 1):
    if ws4[f"D{row}"].value is not None:
        targetToGene[ws4[f"C{row}"].value] = ws4[f"D{row}"].value

s = "甘草、茯苓、白术、半夏、陈皮、鸡内金、党参、麦芽、砂仁、太子参、枳壳、木香、神曲"
s = s.split("、")
for file in s:
    wb = openpyxl.load_workbook(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据" + '\\' + file + ".xlsx")
    ws = wb.worksheets[1]
    wsNew = wb.create_sheet("全部信息")
    count = 2
    wsNew[f"A{1}"] = 'Mol ID'
    wsNew[f"B{1}"] = 'Molecule name'
    wsNew[f"C{1}"] = 'Target name'
    wsNew[f"D{1}"] = 'Gene Symbol'
    for row in range(2, ws.max_row + 1):
        if ws[f"C{row}"].value is not None:
            if ws[f"C{row}"].value in targetToGene:
                wsNew[f"A{count}"].value = ws[f"A{row}"].value
                wsNew[f"B{count}"].value = ws[f"B{row}"].value
                wsNew[f"C{count}"].value = ws[f"C{row}"].value
                wsNew[f"D{count}"].value = targetToGene[ws[f"C{row}"].value]
                count += 1
    wb.save(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据" + '\\' + file + ".xlsx")
