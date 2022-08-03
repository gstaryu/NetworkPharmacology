"""对每味药的Mol ID进行检索，找出那两味药品有相同的Mol ID，并保存在 重复值.xlsx 中，该代码只能检索两味药品是否有相同的Mol ID"""
import openpyxl

s = "甘草、茯苓、白术、半夏、陈皮、鸡内金、党参、麦芽、砂仁、太子参、枳壳、木香"
s = s.split("、")
wb = openpyxl.Workbook()
ws = wb.active
createVar = locals()  # 创建一个字典，用于批量创建列表 https://blog.csdn.net/qq_26144863/article/details/106201021
for file in s:
    wb1 = openpyxl.load_workbook(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据" + '\\' + file + ".xlsx")  # 打开文件
    sheet = wb1.worksheets[2]  # 指定工作表
    createVar[str(file)] = []
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value not in createVar[str(file)]:
            createVar[str(file)].append(sheet.cell(row=row, column=1).value)

count = 1
for i in range(len(s)):
    medicineA = s[i]
    MolAList = createVar[medicineA]
    for j in range(i + 1, len(s)):
        medicineB = s[j]
        MolBList = createVar[medicineB]
        for MolA in MolAList:
            if MolA in MolBList:
                ws.cell(row=count, column=1).value = medicineA
                ws.cell(row=count, column=2).value = medicineB
                ws.cell(row=count, column=3).value = MolA
                count += 1
wb.save(r"C:\Users\guo\Desktop\重复值.xlsx")
