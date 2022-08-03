"""将每味中药的化合物数量进行统计，保存在 化合物数量统计.xlsx 中"""
import openpyxl

wb1 = openpyxl.load_workbook(r"D:\胃癌\创新学习\论文及数据\网络药理学数据\所有药品.xlsx")
ws1 = wb1.worksheets[0]
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2[f"A{1}"] = '中药'
ws2[f"B{1}"] = 'Molecule ID'
ws2[f"C{1}"] = 'Molecule name'
ws2[f"D{1}"] = '靶点数'

count = 2
num = 0
for row in range(2, ws1.max_row + 1):
    num += 1
    if ws1[f"B{row}"].value != ws1[f"B{row + 1}"].value:
        ws2[f"A{count}"] = ws1[f"A{row}"].value
        ws2[f"B{count}"] = ws1[f"B{row}"].value
        ws2[f"C{count}"] = ws1[f"C{row}"].value
        ws2[f"D{count}"] = num
        num = 0
        count += 1

wb2.save(r"D:\胃癌\创新学习\论文及数据\网络药理学数据\化合物数量统计.xlsx")
