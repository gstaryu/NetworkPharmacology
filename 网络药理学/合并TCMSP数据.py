"""将TCMSP中的数据进行合并，保存在 所有药品.xlsx 中"""
import openpyxl

s = "甘草、茯苓、白术、半夏、陈皮、鸡内金、党参、麦芽、砂仁、太子参、枳壳、木香"
s = s.split("、")
wb = openpyxl.Workbook()
ws = wb.active
ws[f'A{1}'].value = "Medicine"
ws[f'B{1}'].value = "Mol ID"
ws[f'C{1}'].value = "Molecule name"
ws[f'D{1}'].value = "Target name"
count = 2
for file in s:
    wb1 = openpyxl.load_workbook(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据" + '\\' + file + ".xlsx")  # 打开文件
    sheet = wb1.worksheets[1]  # 指定工作表
    for row in range(2, sheet.max_row + 1):
        ws.cell(row=count, column=1).value = str(file)
        for col in range(1, 4):
            ws.cell(row=count, column=col + 1).value = sheet.cell(row=row, column=col).value
        count += 1
wb.save(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据" + '\\' + "所有药品.xlsx")
