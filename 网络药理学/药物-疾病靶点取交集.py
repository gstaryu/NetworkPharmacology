"""将核心中药有效成分靶点和稳定逆转疾病靶点取交集，保存到 匹配到的基因.xlsx 中"""
import openpyxl

wb1 = openpyxl.load_workbook(r"D:\胃癌\创新学习\论文及数据\基因数据\GSE79973.xlsx")
ws1 = wb1.worksheets[0]
wb2 = openpyxl.load_workbook(r"D:\胃癌\创新学习\论文及数据\网络药理学数据\药品数据\所有药品（靶点去重）.xlsx")
ws2 = wb2.worksheets[0]
wb3 = openpyxl.load_workbook(r"C:\Users\guo\Desktop\80%稳定逆转基因对.xlsx")
ws3 = wb3.active
wb4 = openpyxl.Workbook()
ws4 = wb4.active

symbol1 = []
for row in range(2, ws1.max_row + 1):
    symbol1.append(ws1[f"A{row}"].value)
symbol2 = []
for row in range(2, ws2.max_row + 1):
    symbol2.append(ws2[f"E{row}"].value)
print(len(set(symbol2)))
symbol3 = []
for row in range(1, ws3.max_row + 1):
    symbol3.append(symbol1[ws3[f"A{row}"].value - 1])
    symbol3.append(symbol1[ws3[f"B{row}"].value - 1])
print(*symbol3,sep='\n')
print(len(symbol3))
print(len(set(symbol3)))
symbol = set(symbol2).intersection(set(symbol3))
symbol = list(symbol)
print(len(symbol))
for row in range(len(symbol)):
    ws4[f"A{row + 1}"].value = symbol[row]
#wb4.save(r"D:\胃癌\创新学习\论文及数据\匹配到的基因.xlsx")
