"""将所有药品中的蛋白质名称（Target name）转换成Gene Symbol，并写入 所有药品2（蛋白质名称转基因名称）.xlsx 文件中，使用生信自学网的uniport数据"""
import openpyxl
import pprint  # pprint是一个打印函数，用于打印字典，列表，元组等对象的内容
import winsound

wbUniport = openpyxl.load_workbook(r"C:\Users\guo\Desktop\uniport.xlsx")  # 打开文件
wsUniport = wbUniport.worksheets[0]  # 指定工作表
wbMedicine = openpyxl.load_workbook(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据\所有药品.xlsx")
wsMedicine = wbMedicine.worksheets[0]
wbTransform = openpyxl.Workbook()  # 新建一个文件
wsTransform = wbTransform.active

wsTransform[f"A{1}"] = 'Medicine'
wsTransform[f"B{1}"] = 'Mol ID'
wsTransform[f"C{1}"] = 'Molecule name'
wsTransform[f"D{1}"] = 'Target name'
wsTransform[f"E{1}"] = 'Gene Symbol'

targetToGene = {}
for row in range(2, wsUniport.max_row + 1):
    targetToGene[wsUniport.cell(row=row, column=4).value] = str(wsUniport.cell(row=row, column=5).value)
print('写入pprint数据中...')  # 以.py文件的形式保存字典，便于之后使用
resultFile = open('targetToGene.py', 'w', encoding='utf-8')
resultFile.write('targetToGene = ' + pprint.pformat(targetToGene))
count = 0
for row in range(2, wsMedicine.max_row + 1):
    wsTransform[f"A{row}"].value = wsMedicine[f"A{row}"].value
    wsTransform[f"B{row}"].value = wsMedicine[f"B{row}"].value
    wsTransform[f"C{row}"].value = wsMedicine[f"C{row}"].value
    wsTransform[f"D{row}"].value = wsMedicine[f"D{row}"].value
    if wsMedicine.cell(row=row, column=4).value in targetToGene:
        gene = targetToGene[wsMedicine.cell(row=row, column=4).value].split()
        wsTransform.cell(row=row, column=5).value = gene[0]
    else:
        count += 1
        continue
print(f'未找到的蛋白质数量：{count}, 占比：{count / (wsMedicine.max_row - 1)}')
wbTransform.save(r"C:\Users\guo\Desktop\胃癌\创新学习\论文及数据\网络药理学数据\所有药品2（蛋白质名称转基因名称）.xlsx")
winsound.Beep(1000, 1000)  # 运行结束，发出提示音
